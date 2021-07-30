import datetime
import os.path
from random import randint

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

path = os.path.join(os.curdir, "files")
name = 'products_selected.xlsx'

required_to_fill = ''

categories = ('bool_values', 'CMFR', 'RiskCategory', 'Country', 'BrandGroup', 'TransportConditions', 'ProductUnit',
              'ProductPackingType', 'FNS', 'TNVED', 'OKPD', 'VAT', 'packagepack_type', 'packagepack_material',)


def to_xlsx(file):
    pass


def to_xls(file):
    pass


class Xls:
    def __init__(self, path: str = '', name: str = '', required: list = None):
        self.__file = os.path.join(path, name)
        self.__sheet = None
        self.__cell = None
        self.__required = required
        self.__rows = []

    @property
    def sheet(self):
        self.__sheet = openpyxl.open(self.__file)['ФПИ']
        return self

    def cell(self, cell: str = 'A1'):
        self.__cell = self.__sheet[cell]
        return self.__cell

    def find(self, value, sheet: Worksheet, from_row: int = 1, col: int = 1) -> tuple[int, int]:
        for row in range(1, sheet.max_row):
            if sheet.cell(row, col).value == value:
                return row, col
        return sheet.max_row, col

    def _find_row_or_throw_exception(self, sheet: Worksheet, col: int = 2, value: str = '') -> int:
        for row in range(1, sheet.max_row):
            if sheet.cell(row, col).value == value:
                return row
        raise Exception(f"Cell with value '{value}' doesn't exist in sheet {sheet.title} in column #{col}")

    def _count_of_occurences(self, sheet_name: str, start_row: int = 1, col: int = 1, value: str = '') -> int:
        count = 0
        try:
            sheet = openpyxl.open(self.__file)[sheet_name]
        except:
            raise Exception(f"Worksheet with name '{sheet_name}' doesn't exist")
        for row in range(start_row, sheet.max_row):
            if sheet.cell(row, col).value == value:
                count += 1
            else:
                return count
        raise Exception(f"Cell with value '{value}' doesn't exist in sheet '{sheet_name}' in column #{col}")

    @property
    def color(self) -> str:
        return self.__cell.fill.fgColor.rgb

    @property
    def required(self):
        if self.__required is None:
            self.__required = ['A1']
        return self.__required

    def row(self, index: int = None):
        rows = self.__sheet.rows
        for i in range(index):
            self.__rows.append(next(rows))
        return [[i.value for i in self.__rows[j]] for j in range(len(self.__rows))]

    def product_category(self, vals: list):
        valid_product_group = []
        sheets = ('categories_1', 'categories_2', 'categories_3', 'categories_4')
        sheet = openpyxl.open(self.__file)['categories']
        try:
            row = self._find_row_or_throw_exception(sheet, col=1, value=vals[0])
            valid_product_group.append(sheet.cell(row, 1).value)
        except:
            row = randint(1, sheet.max_row + 1)
            valid_product_group.append(sheet.cell(row, 1).value)
        vals = vals[1:]
        for i in range(len(sheets)+1):
            sheet = openpyxl.open(self.__file)[sheets[i]]
            try:
                row = self._find_row_or_throw_exception(sheet, value=vals[i])
                valid_product_group.append(sheet.cell(row, 2).value)
            except:
                try:
                    f = self._find_row_or_throw_exception(sheet, col=1, value=valid_product_group[-1])
                    c = self._count_of_occurences(sheet, f, value=valid_product_group[-1])
                    row = randint(f, f + c)
                    valid_product_group.append(sheet.cell(row, 2).value)
                except:
                    return valid_product_group
        return valid_product_group

    def dropdown(self, category, next_category, col: int = 1, rowx: int = 1, value: str = None):
        sheet = openpyxl.open(self.__file)['catalogs']
        start = self.find(category, sheet)[0]
        end = self.find(next_category, sheet, start)[0]
        if category is None:
            rowx = randint(start, end)
        elif category is not None:
            rowx = self.find(value, sheet)[0]
            if rowx < start or rowx > end:
                rowx = randint(start, end)
        return sheet.cell(rowx, col)


class Milk:
    def __init__(self):
        self._table = Xls(path, name, required=['A1', 'B2'])

    def _set_product_category(self, cells: list):
        product_category = self._table.product_category(['молоко и молочные товары'])
        sheet = self._table.sheet
        for i in range(len(product_category)):
            sheet.cell(cells[i]).value = product_category[i]

    def _set_value(self, value, _cell):
        self._table.sheet.cell(_cell).value = value

    def _set_random_value_from_list(self, values, _cell):
        self._table.sheet.cell(_cell).value = values[randint(0, len(values))]

    def table(self):
        now = datetime.datetime.today().strftime('%d %b %Y %X')
        product = f'Тест молоко at {now}'
        milk._set_product_category(cells=['E5', 'F5', 'G5', 'H5', 'I5'])
        milk._set_value(product, "C5")


def test_open():
    file = Xls(path, name)
    sheet = file.sheet
    # return sheet.cell('E5').value
    # return sheet.row(5)
    # todo написать функции для ввода в нужную ячейку а также алгоритмы для базовы случаев- общий случай, молочка, алко и товары животного происхождения
    # return file.dropdown(categories[4], categories[5], value='Продукция низкого риска').value
    return file.product_category(['молоко и молочные товары', 'молоко'])


if __name__ == '__main__':
    milk = Milk()
    milk.table()

